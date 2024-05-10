import time
import os
import allure
import pytest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (ElementNotVisibleException, ElementNotSelectableException)

# Data Driven Test case for the VWO Login Page.
# Valid and Invalid Logins for the VWO login page.

def read_credentials_from_excel(file_path):
    credentials = []
    workbook = openpyxl.load_workbook(file_path)#The load_workbook() function opens file for reading
    sheet = workbook.active#The object of the workbook.active has been created in the script to read the values of the max_row and the max_column properties.
    for row in sheet.iter_rows(min_row=2, values_only=True): #for loop to read each row from the excel sheet
        username1, password1 = row
        credentials.append({
            "username" : username1,
            "password" : password1
        })
    print(credentials)
    return credentials

file_path_fromos = os.getcwd() + "/py2xtestdata.xlsx"
print(file_path_fromos)

#instead of for loop to read each row from excel sheet, we use parametrize to read all data(each row)
#Parameterizing of a test is done to run the test against multiple sets of inputs.
@pytest.mark.parametrize("user_cred", read_credentials_from_excel(file_path_fromos))
@allure.title("Verify the Invalid Login with the Excel Testdata.")
@allure.description("TC#1 - 10 Invalid login verification for app.vwo.com")
def test_vwo_login(user_cred):
    username2 = user_cred["username"]
    password2 = user_cred["password"]
    print(username2, password2)
    vwo_login(username=username2, password=password2)

def vwo_login(username,password):
    driver = webdriver.Chrome()
    driver.get("https://app.vwo.com")

    email_input = driver.find_element(By.CSS_SELECTOR, "[name='username']")
    pswd_input = driver.find_element(By.CSS_SELECTOR, "[name='password']")
    email_input.send_keys(username)
    pswd_input.send_keys(password)

    submit_btn = driver.find_element(By.XPATH, "//button[@id='js-login-btn']")
    submit_btn.click()

    time.sleep(5)

    result = driver.current_url
    print(result)
    if result != "https://app.vwo.com/#/dashboard" :
        ignore_list = [ElementNotVisibleException, ElementNotSelectableException]
        wait = WebDriverWait(driver,timeout=20, poll_frequency=1, ignored_exceptions=ignore_list)
        wait.until(
            EC.visibility_of_element_located((By.ID, "js-notification-box-msg"))
        )
        error_msg = driver.find_element(By.ID, "js-notification-box-msg")
        assert error_msg.text == "Your email, password, IP address or location did not match"
    else :
        wait = WebDriverWait(driver=driver, timeout=10)
        wait.until(
            EC.text_to_be_present_in_element((By.CSS_SELECTOR, ".page-heading"),"Dashboard")
        )
        heading_element = driver.find_element(By.XPATH, "//span[@data-qa='lufexuloga']")
        assert heading_element.text == "Py2xATB"
    driver.quit()